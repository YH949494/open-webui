"""
title: Excel Analyzer
author: open-webui
version: 1.1.0
required_open_webui_version: 0.5.0
requirements: openpyxl
description: >
  Inspect spreadsheet files (.xlsx/.xlsm/.xls/.csv) that a user has attached to
  the chat. The tool resolves the *actual server-side path* of the uploaded
  file and reads it directly with openpyxl/pandas instead of relying on RAG
  extracted text. This is the companion tool to the "skip RAG for spreadsheets"
  upload path: spreadsheets are stored as raw data files (never chunked into the
  vector DB and never injected into the prompt), so a tool must open the file on
  disk to analyze it.

INSTALLATION
------------
This file is the source for a *custom* Open WebUI tool. Built-in tools are
registered in ``open_webui/tools/builtin.py``; this module is intentionally NOT
auto-loaded. To use it:

  Workspace -> Tools -> (+) -> paste the contents of the ``Tools`` class below
  (or this whole file) -> Save.

Then attach a spreadsheet to a chat and ask, e.g.:

  "Use the Excel analyzer tool to inspect the uploaded file.
   Return only sheet names and row counts."

HOW FILE ACCESS WORKS
---------------------
Open WebUI passes attached files to a tool through the reserved ``__files__``
parameter (``metadata['files']``). Each entry looks like::

    {"type": "file", "id": "<file_id>", "name": "Foo.xlsx",
     "file": { ...full File record incl. "path"... }}

The stored ``file.path`` is a *storage URI* (``s3://``, ``gs://``, an Azure
URL, or — for local/Fly persistent-volume storage — an absolute path). It is
NOT necessarily openable with ``open()``. ``resolve_uploaded_file_path`` turns a
``file_id`` into a readable local path: it returns the absolute path as-is for
local storage, and downloads cloud objects into ``UPLOAD_DIR`` for cloud
providers. The same code path works on a Fly.io persistent volume because
``UPLOAD_DIR`` is the mounted volume there.
"""

from __future__ import annotations

import json
import logging
import os

from pydantic import BaseModel, Field

log = logging.getLogger(__name__)

# Extensions this tool knows how to open directly.
_SPREADSHEET_EXTENSIONS = {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls', '.csv'}
_MAX_COL_NAMES = 20
_MAX_INSIGHTS = 10
_MAX_SAMPLE_ROWS = 5000
_OUTPUT_SIZE_LIMIT = 12000


class Tools:
    class Valves(BaseModel):
        max_files: int = Field(
            default=10,
            description='Maximum number of attached spreadsheet files to inspect in one call.',
        )
        expose_paths_in_output: bool = Field(
            default=False,
            description=(
                'If true, include the resolved absolute server path in the tool output. '
                'Leave false in production so sensitive paths stay in the server logs only.'
            ),
        )
        lightweight_mode: bool = Field(
            default=True,
            description=(
                'When true (default), return only compact metadata: sheet names, row/column counts, '
                'and up to 20 column names. No data previews or full dataframe reads.'
            ),
        )

    def __init__(self):
        self.valves = self.Valves()
        # We read the raw uploaded file ourselves; we do NOT want Open WebUI to
        # strip the files from the request or treat this as a citation handler.
        self.file_handler = False
        self.citation = False

    async def inspect_uploaded_spreadsheet(
        self,
        file_id: str = '',
        include_insights: bool = False,
        __files__: list | None = None,
        __user__: dict | None = None,
    ) -> str:
        """
        Inspect an uploaded spreadsheet (.xlsx/.xlsm/.xls/.csv) and return sheet
        names, row/column counts, and up to 20 column headers per sheet.
        Optionally include up to 10 structural insights. Output is always compact
        JSON capped at 12 000 characters — no dataframe previews, no narrative.

        :param file_id: Optional id of a specific attached file. If omitted, the most recent spreadsheet is used.
        :param include_insights: If true, include up to 10 structural findings per sheet.
        :return: Compact JSON with sheet metadata.
        """
        user = await self._resolve_user(__user__)
        if user is None:
            return json.dumps({'error': 'User context not available; cannot access uploaded files.'})

        candidates = self._collect_candidates(__files__, file_id)
        if not candidates:
            return json.dumps(
                {
                    'error': 'No uploaded file was available to the tool.',
                    'hint': 'Attach a spreadsheet to the chat before invoking this tool.',
                    'received_files': self._describe_files(__files__),
                }
            )

        from open_webui.utils.files import resolve_uploaded_file_path

        results = []
        for item in candidates[: max(1, self.valves.max_files)]:
            fid = item.get('id')
            fname = item.get('name') or item.get('filename') or (item.get('file') or {}).get('filename')
            # Pass the storage path from the attachment dict so the resolver can
            # use it directly without a DB round-trip.
            attachment_path = (item.get('file') or {}).get('path') or None

            local_path = await resolve_uploaded_file_path(fid, user=user, attachment_path=attachment_path)
            if not local_path:
                results.append(self._missing_path_diagnostic(item, fid, fname))
                continue

            log.info(f'inspect_uploaded_spreadsheet: analyzing file_id={fid} at {local_path}')
            analysis = self._analyze(local_path, fname, include_insights=include_insights)
            if self.valves.expose_paths_in_output:
                analysis['resolved_path'] = local_path
            results.append(analysis)

        return self._guard_output({'files': results})

    # --- helpers -----------------------------------------------------------

    async def _resolve_user(self, __user__: dict | None):
        """Turn the ``__user__`` dict Open WebUI injects into a full user object."""
        if not __user__ or not __user__.get('id'):
            return None
        try:
            from open_webui.models.users import Users

            return await Users.get_user_by_id(__user__['id'])
        except Exception as e:
            log.warning(f'inspect_uploaded_spreadsheet: could not resolve user: {e}')
            return None

    def _collect_candidates(self, __files__: list | None, file_id: str) -> list:
        """Pick which attached file(s) to inspect."""
        files = [f for f in (__files__ or []) if isinstance(f, dict)]

        if file_id:
            for f in files:
                if f.get('id') == file_id:
                    return [f]
            return [{'id': file_id}]

        spreadsheets = [f for f in files if self._is_spreadsheet(f)]
        return spreadsheets or files

    @staticmethod
    def _is_spreadsheet(item: dict) -> bool:
        name = item.get('name') or item.get('filename') or (item.get('file') or {}).get('filename') or ''
        return os.path.splitext(name.lower())[1] in _SPREADSHEET_EXTENSIONS

    @staticmethod
    def _describe_files(__files__: list | None) -> list:
        described = []
        for f in __files__ or []:
            if isinstance(f, dict):
                described.append(
                    {
                        'id': f.get('id'),
                        'name': f.get('name') or f.get('filename'),
                        'type': f.get('type'),
                        'keys': sorted(f.keys()),
                    }
                )
        return described

    @staticmethod
    def _missing_path_diagnostic(item: dict, fid, fname) -> dict:
        file_meta = (item.get('file') or {}).get('meta') or {}
        return {
            'error': 'file not found; the file path was not accessible to the tool',
            'file_id': fid,
            'filename': fname,
            'available_item_keys': sorted(item.keys()),
            'available_meta_keys': sorted(file_meta.keys()),
        }

    @staticmethod
    def _guard_output(payload: dict) -> str:
        """Ensure output stays under _OUTPUT_SIZE_LIMIT characters with valid JSON."""
        raw = json.dumps(payload, ensure_ascii=False, separators=(',', ':'))
        if len(raw) <= _OUTPUT_SIZE_LIMIT:
            return raw

        # Pass 1: drop insights
        for f in payload.get('files', []):
            for s in f.get('sheets', []):
                s.pop('insights', None)
        payload['truncated'] = True
        raw = json.dumps(payload, ensure_ascii=False, separators=(',', ':'))
        if len(raw) <= _OUTPUT_SIZE_LIMIT:
            return raw

        # Pass 2: drop column_names
        for f in payload.get('files', []):
            for s in f.get('sheets', []):
                s.pop('column_names', None)
        raw = json.dumps(payload, ensure_ascii=False, separators=(',', ':'))
        if len(raw) <= _OUTPUT_SIZE_LIMIT:
            return raw

        # Pass 3: cap each file to 10 sheets
        for f in payload.get('files', []):
            if len(f.get('sheets', [])) > 10:
                f['sheets'] = f['sheets'][:10]
                f['sheets_truncated'] = True
        return json.dumps(payload, ensure_ascii=False, separators=(',', ':'))

    @staticmethod
    def _analyze(local_path: str, fname: str | None, include_insights: bool = False) -> dict:
        """Dispatch to the correct reader based on file extension."""
        ext = os.path.splitext((fname or local_path).lower())[1]
        display_name = fname or os.path.basename(local_path)

        try:
            if ext == '.csv':
                return Tools._analyze_csv(local_path, display_name, include_insights)
            if ext == '.xls':
                return Tools._analyze_xls(local_path, display_name, include_insights)
            # Default: xlsx / xlsm / xltx / xltm via openpyxl streaming read.
            return Tools._analyze_xlsx(local_path, display_name, include_insights)
        except Exception as e:
            log.exception(f'inspect_uploaded_spreadsheet: failed to read {local_path}: {e}')
            return {'filename': display_name, 'error': f'Failed to read spreadsheet: {e}'}

    @staticmethod
    def _analyze_xlsx(local_path: str, display_name: str, include_insights: bool) -> dict:
        """Read xlsx metadata via openpyxl without loading cell data into memory."""
        import openpyxl

        wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
        try:
            sheets = []
            for ws in wb.worksheets:
                # max_row/max_column from the workbook manifest — no full read.
                row_count = int(ws.max_row) if ws.max_row is not None else 0
                col_count = int(ws.max_column) if ws.max_column is not None else 0

                col_names: list[str] = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    col_names = [str(c) if c is not None else '' for c in row[:_MAX_COL_NAMES]]
                    break

                sheet_info: dict = {
                    'sheet_name': ws.title,
                    'rows': row_count,
                    'columns': col_count,
                    'column_names': col_names,
                }
                if include_insights:
                    sheet_info['insights'] = Tools._structural_insights(
                        row_count, col_count, col_names
                    )
                sheets.append(sheet_info)
            return {'filename': display_name, 'sheets': sheets}
        finally:
            wb.close()

    @staticmethod
    def _analyze_xls(local_path: str, display_name: str, include_insights: bool) -> dict:
        """Read legacy .xls metadata via xlrd (no pandas full-sheet load)."""
        try:
            import xlrd

            wb = xlrd.open_workbook(local_path)
            sheets = []
            for sh in wb.sheets():
                nrows = int(sh.nrows)
                ncols = int(sh.ncols)
                col_names = (
                    [str(sh.cell_value(0, c)) for c in range(min(ncols, _MAX_COL_NAMES))]
                    if nrows > 0
                    else []
                )
                sheet_info: dict = {
                    'sheet_name': sh.name,
                    'rows': nrows,
                    'columns': ncols,
                    'column_names': col_names,
                }
                if include_insights:
                    sheet_info['insights'] = Tools._structural_insights(nrows, ncols, col_names)
                sheets.append(sheet_info)
            return {'filename': display_name, 'sheets': sheets}
        except ImportError:
            # xlrd not installed — fall back to pandas header-only read.
            import pandas as pd

            header_frames = pd.read_excel(local_path, sheet_name=None, header=0, nrows=0)
            sheets = []
            for name, df in header_frames.items():
                col_names = [str(c) for c in df.columns[:_MAX_COL_NAMES]]
                sheet_info = {
                    'sheet_name': name,
                    'rows': None,
                    'columns': len(df.columns),
                    'column_names': col_names,
                    'note': 'row count unavailable without xlrd',
                }
                if include_insights:
                    sheet_info['insights'] = Tools._structural_insights(None, len(df.columns), col_names)
                sheets.append(sheet_info)
            return {'filename': display_name, 'sheets': sheets}

    @staticmethod
    def _analyze_csv(local_path: str, display_name: str, include_insights: bool) -> dict:
        """Count CSV rows by line scan and read only the header with pandas."""
        import pandas as pd

        row_count = 0
        try:
            with open(local_path, 'rb') as fh:
                row_count = max(0, sum(1 for _ in fh) - 1)  # subtract header line
        except Exception:
            pass

        header_df = pd.read_csv(local_path, nrows=0)
        col_count = len(header_df.columns)
        col_names = [str(c) for c in header_df.columns[:_MAX_COL_NAMES]]

        sheet_info: dict = {
            'sheet_name': 'Sheet1',
            'rows': row_count,
            'columns': col_count,
            'column_names': col_names,
        }
        if include_insights:
            sheet_info['insights'] = Tools._structural_insights(row_count, col_count, col_names)
        return {'filename': display_name, 'sheets': [sheet_info]}

    @staticmethod
    def _structural_insights(rows, columns, col_names: list[str]) -> list[str]:
        """Return up to _MAX_INSIGHTS compact structural observations."""
        findings: list[str] = []
        if rows is not None:
            findings.append(f'{rows} data rows, {columns} columns')
            if rows > _MAX_SAMPLE_ROWS:
                findings.append(f'Large sheet (>{_MAX_SAMPLE_ROWS} rows); sample for analysis')
            elif rows == 0:
                findings.append('Sheet appears empty')
        else:
            findings.append(f'{columns} columns (row count not available)')
        if col_names:
            extra = columns - len(col_names) if columns > len(col_names) else 0
            summary = ', '.join(col_names)
            if extra:
                summary += f' … +{extra} more'
            findings.append(f'Columns: {summary}')
        return findings[:_MAX_INSIGHTS]
